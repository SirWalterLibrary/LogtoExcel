# import modules
import csv
import pandas as pd
from sys import exit
from tkinter import filedialog
from openpyxl.worksheet import cell_range 
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

class Data:

    def __init__(self,index=None):
        # create empty lists
        self.data = []
        self.headers = getHeaders("headers.txt",index)
        
    def __call__(self,other):
        # get initial headers and its range
        initial_headers = getHeaders("headers.txt",0)
        initial_range = [x + 1 for x in list(range(len(initial_headers)))]
        
        # split data to its respective category (i.e. Dimensions, Contour Verify, Corner)
        self.range = initial_range + [index for index, item in enumerate(other.headers) if item in self.headers]
        self.data = (other.data[self.range]).set_axis((initial_headers + self.headers),axis=1)

    def getData(self):
        # open dialog to browse file in File Explorer 
        filepath = filedialog.askopenfilename(initialdir="/", 
                                              title="Select a File", 
                                              filetypes=[("Log files", ".txt .log")])
        
        # check if user imported file 
        if filepath != '':
            # append file contents into data list 
            with open(filepath) as logFile:
                for row in csv.reader(logFile, delimiter=';'):
                    self.data.append(row)
        else:
            # error if no file was imported
            print("No log file was imported")
            exit()

    def parseData(self):
        # create data frame and parse through data
        self.getData()
        self.data = pd.DataFrame(self.data).iloc[:,1:-1].dropna()

    def paste2Excel(self,worksheet,title):
        # title the worksheet
        worksheet.title = title

        # paste data frame into Excel worksheet
        for rows in dataframe_to_rows(self.data, index=False, header=True):
            worksheet.append(rows)

        # format table from data
        formatTable(worksheet)

def formatTable(worksheet):
    # define range
    full_range = cell_range.CellRange(min_col=worksheet.min_column,
                                      min_row=worksheet.min_row,
                                      max_col=worksheet.max_column,
                                      max_row=worksheet.max_row).coord

    # set table format
    mediumStyle =TableStyleInfo(name='TableStyleMedium1',
                                showRowStripes=True)
    # create a table
    table = Table(ref=full_range,
                  displayName=(worksheet.title).replace(" ", "_"),
                  tableStyleInfo=mediumStyle)

    # add the table to the worksheet
    worksheet.add_table(table)

def getHeaders(filename, line_number):
    # open header file
    with open(filename) as file:
        # get all headers
        if line_number == None:
            return [x.strip(' ') for x in (" ".join(line.strip() for line in file)).split(';')]
        else:
            for i, line in enumerate(file):
                # get only the headers from the line
                if i == line_number:
                    return list(filter(None,(line.strip()).split(';')))
