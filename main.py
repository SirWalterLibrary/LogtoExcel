# summary: main.py converts log to an Excel speadsheet formatted for analysis
 
# import required modules
import sys, os, csv, shutil
from sys import exit
from os.path import exists
import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook 
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.cell_range import CellRange 
from openpyxl.worksheet.table import Table, TableStyleInfo

def resource_path(relative_path):

    # get absolute path to resource
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))

    # return the path for the python executable
    return os.path.join(base_path, relative_path)

def download_file():

    # let the user download a copy of the Expression Editor xml
    src = r"C:/Users/mohammo/Documents/Projects/Dims Template Code/Python/LogtoExcel/Dimension_Expression_Editor.xml"
    des = filedialog.asksaveasfilename(title="Download", initialdir="Downloads", initialfile="Dimension_Expression_Editor.xml",  filetypes=[("Expression Editor",".xml")])

    # return the copied file
    return shutil.copyfile(src,des)

def run(len_entry,wid_entry,hei_entry,window,import_entry,export_entry):

    def script():
        
        # retrieve excel file path
        excel_file = export_entry.get()

        class Log:
            def __init__(self):
                
                # create empty data list
                self.data = []

                # get headers from headers.txt
                self.headers = self.getHeaders()
                
            def getData(self):
                
                # read log file & split the data
                with open(log_path) as logFile:
                    for row in csv.reader(logFile, delimiter=';'):
                        self.data.append(row)

            def parseData(self):
                
                # create data frame & parse through data
                self.getData()
                self.data = pd.DataFrame(self.data).iloc[:,1:-1].dropna()

            def getHeaders(self, line_number=None):
                
                # specify header file
                filename = resource_path("headers.txt")

                # open header file & get all the headers
                with open(filename) as file:

                    if line_number is None:
                        return [x.strip(' ') for x in (" ".join(line.strip() for line in file)).split(';')]
                    else:
                        for i, line in enumerate(file):
                            # get only the headers from the line
                            if i == line_number:
                                return list(filter(None,(line.strip()).split(';')))

        class Category(Log):

            def __init__(self, index=None):
                
                # calls original __init__()
                super().__init__()
                self.headers = self.getHeaders(index)

            def __call__(self,other):
                
                # get initial headers & its range
                initial_headers = self.getHeaders(0)
                initial_range = [x + 1 for x, header in enumerate(initial_headers)]
                
                # split data to its respective category (i.e. Dimensions, Contour Verify, Corner)
                self.range = initial_range + [index + 1 for index, item in enumerate(other.headers) if item in self.headers]
                self.data = (other.data[self.range]).set_axis((initial_headers + self.headers),axis=1)

            def formatTable(self, worksheet):
                
                # define the range of cells
                full_range = CellRange(min_col=worksheet.min_column,
                                        min_row=worksheet.min_row,
                                        max_col=worksheet.max_column,
                                        max_row=worksheet.max_row).coord

                # set table format
                mediumStyle =TableStyleInfo(name='TableStyleMedium1',
                                            showRowStripes=True)
                # create a table
                table = Table(ref=full_range,
                                displayName=(worksheet.title).replace(" ", "_"),
                                tableStyleInfo=mediumStyle)

                # add the table to the worksheet
                worksheet.add_table(table)

            def paste2Excel(self,worksheet,title,add_cols=None):
                
                # title the worksheet
                worksheet.title = title
                
                # add occupied columns
                if not add_cols == None:
                    content = [''] * len(add_cols)
                    self.data[add_cols] = content 

                # paste the data frame into an Excel worksheet
                for rows in dataframe_to_rows(self.data, index=False, header=True):
                    worksheet.append(rows)

                # format table from data
                self.formatTable(worksheet)
        
        # initialize log
        log = Log()

        # parse data from log
        log.parseData()

        # initialize "Results" category
        res = Category(1)
        res(log)

        # initialize "Dimensions" category
        dim = Category(2)
        dim(log)

        # initialize "Contour Verify" category
        c_v = Category(3)
        c_v(log)

        # open a workbook
        wb = Workbook()

        # create "Results" sheet & paste data
        ws1 = wb.active
        expected_cols = ['Expected L','Expected W','Expected H']
        error_cols = ['Error L','Error W','Error H']
        res.paste2Excel(ws1,"Results", expected_cols + error_cols)

        # create a formula to find difference between measured & expected dimensions
        for row_num in range(2, ws1.max_row+1):
            ws1['T'+ str(row_num)] = '=IF(NOT(ISNUMBER(VALUE(LEFT(Q'+str(row_num)+',1)))),"",F'+str (row_num)+'-Q'+str (row_num)+')'
            ws1['U'+ str(row_num)] = '=IF(NOT(ISNUMBER(VALUE(LEFT(R'+str(row_num)+',1)))),"",G'+str (row_num)+'-R'+str (row_num)+')'
            ws1['V'+ str(row_num)] = '=IF(NOT(ISNUMBER(VALUE(LEFT(S'+str(row_num)+',1)))),"",H'+str (row_num)+'-S'+str (row_num)+')'
        
        # color cells red if error is out of spec
            red_color = 'ffc7ce'
            red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
            ws1.conditional_formatting.add('T2:T'+str(ws1.max_row), CellIsRule(operator='notBetween', formula=['-' + length,length], fill=red_fill))
            ws1.conditional_formatting.add('U2:U'+str(ws1.max_row), CellIsRule(operator='notBetween', formula=['-' + width,width], fill=red_fill))
            ws1.conditional_formatting.add('V2:V'+str(ws1.max_row), CellIsRule(operator='notBetween', formula=['-' + height,height], fill=red_fill))

        # create "Contour Verify" sheet & paste data
        ws2 = wb.create_sheet()
        dim.paste2Excel(ws2,"Contour Verify")

        # create "Corner" sheet & paste data
        ws3 = wb.create_sheet()
        c_v.paste2Excel(ws3,"CCorner")

        # save workbook
        wb.save(excel_file)

        # check if excel was created from log
        if exists(excel_file):
            messagebox.showinfo(title="Success!", message="Log file successfully parsed to Excel!")
            os.startfile(excel_file)
        else:
            print("ERROR: Script did not run successfully to Excel...")

    length = len_entry.get()
    width = wid_entry.get()
    height = hei_entry.get()

    # ensure that user inputs tolerances
    try:
        int(length), int(width), int(height)
    except:
        messagebox.showerror("Error", "Please input numeric tolerances")
        return

    # check if the log file exists
    log_path = import_entry.get()
    
    if not os.path.exists(log_path):
        messagebox.showerror("Error","You must input an existing log file")
        return
    
    # attempt to run the script, otherwise output error message
    try:
        script()
    except:
        messagebox.showerror("Error", "Script cannot parse data. Check if log is formatted correctly!")
        return
    
    # close the UI & python script
    window.destroy()
    exit(0)

def import_file(import_entry):

    # prompt user to input log file
    log_path = filedialog.askopenfilename(title="Select a File",filetypes=[("Log files", ".txt .log")])
    
    # populate log path to UI entry
    import_entry.delete(0,"end")
    import_entry.insert(0,log_path)

    return

def export_file(export_entry):

    # prompt user to specify output location for excel file
    excel_path = filedialog.asksaveasfilename(title="Save as", filetypes=[("Excel File",".xlsx")], defaultextension=".xlsx")
    
    # populate excel path to UI entry
    export_entry.delete(0,"end")
    export_entry.insert(0,excel_path)

    return
    
def main():

    # create a UI window
    window = tk.Tk()
    window.title("Input Log Data")
    window.geometry("600x300")
    frame = tk.Frame(window)
    frame.pack()

    # declare an "Input Tolerance" frame
    input_tol_frame = tk.LabelFrame(frame,text="Input Tolerance")
    input_tol_frame.grid(column=1, sticky="ns")

    # create an entry for "Length"
    len_label = tk.Label(input_tol_frame,text="Length")
    len_label.grid(row=0,column=0)
    len_entry = tk.Entry(input_tol_frame, width=15)
    len_entry.grid(row=1, column=0)

    # create an entry for "Width"
    wid_label = tk.Label(input_tol_frame,text="Width")
    wid_label.grid(row=2, column=0)
    wid_entry = tk.Entry(input_tol_frame, width=15)
    wid_entry.grid(row=3, column=0)

    # create an entry for "Height"
    hei_label = tk.Label(input_tol_frame,text="Height")
    hei_label.grid(row=4,column=0)
    hei_entry = tk.Entry(input_tol_frame, width=15)
    hei_entry.grid(row=5,column=0)

    # declare an "Input File" frame
    file_frame = tk.LabelFrame(frame, text="Import/Export")
    file_frame.grid(row=0, column=0)

    # create an "Input File" button
    import_entry = tk.Entry(file_frame, width=50)
    import_entry.grid(row=0, column=1)
    import_button = tk.Button(file_frame, text="Import file", width= 8, command=lambda:import_file(import_entry))
    import_button.grid(row=0, column=0)

    # create an "Output Excel" button
    export_entry = tk.Entry(file_frame, width=50)
    export_entry.grid(row=1, column=1)
    export_button = tk.Button(file_frame, text="Save as", width= 8, command=lambda:export_file(export_entry))
    export_button.grid(row=1, column=0)

    # create a "Download Expression Editor" button
    download_button = tk.Button(file_frame, text="Download", width= 50, command=lambda:download_file())
    download_button.grid(row=2, column=1)

    # declare a "Run with" frame
    validate_frame = tk.LabelFrame(frame, text="Run the script?")
    validate_frame.grid(row=3, column=0)

    # create a "Run" button 
    with_button = tk.Button(validate_frame, text="Run", width=40, command=lambda:run(len_entry,wid_entry,hei_entry,window,import_entry,export_entry))
    with_button.grid(row=0, column=0)

    # adjust padding
    for widget in input_tol_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    for widget in file_frame.winfo_children():
        widget.grid_configure(padx=10, pady=10)

    for widget in validate_frame.winfo_children():
        widget.grid_configure(padx=30, pady=10)

    # end code with loop
    window.mainloop()

if __name__ == "__main__":
    main()