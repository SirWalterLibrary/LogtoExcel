# summary: main.py gets dim log and outputs an Excel table formatted for analysis
 
# import modules
import sys, os, csv
from sys import exit
from os.path import exists
import pandas as pd
import tkinter as tk
from tkinter import messagebox, Tk, filedialog
from openpyxl import Workbook 
from openpyxl.worksheet.cell_range import CellRange 
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

def main():

    def script(filepath):
        class Log:
            def __init__(self):
                # create empty data list
                self.data = []

                # get headers from headers.txt
                self.headers = self.getHeaders()

            def getData(self):
                # replace default icon to log.ico 
                win = Tk()
                win.withdraw()
                win.iconbitmap(r'src/log.ico')

                # open dialog to browse file in File Explorer
                #filepath = filedialog.askopenfilename(initialdir='/', title="Select a File", filetypes=[("Log files", ".txt .log")])
                
                # check if user imported file 
                if filepath != '':
                    # append file contents into data list 
                    with open(filepath) as logFile:
                        for row in csv.reader(logFile, delimiter=';'):
                            self.data.append(row)
                else:
                    # error if no file was imported
                    print("No log file was imported")
                    exit()

            def parseData(self):
                # create data frame and parse through data
                self.getData()
                self.data = pd.DataFrame(self.data).iloc[:,1:-1].dropna()


            def getHeaders(self, line_number=None):
                # specify header file
                filename = 'src/headers.txt'

                # open header file
                with open(filename) as file:
                    # get all headers
                    if line_number is None:
                        return [x.strip(' ') for x in (" ".join(line.strip() for line in file)).split(';')]
                    else:
                        for i, line in enumerate(file):
                            # get only the headers from the line
                            if i == line_number:
                                return list(filter(None,(line.strip()).split(';')))

        class Category(Log):

            def __init__(self, index=None):
                # calls original __init__()
                super().__init__()
                self.headers = self.getHeaders(index)

            def __call__(self,other):
                # get initial headers and its range
                initial_headers = self.getHeaders(0)
                initial_range = [x + 1 for x, header in enumerate(initial_headers)]
                
                # split data to its respective category (i.e. Dimensions, Contour Verify, Corner)
                self.range = initial_range + [index + 1 for index, item in enumerate(other.headers) if item in self.headers]
                self.data = (other.data[self.range]).set_axis((initial_headers + self.headers),axis=1)
            
            def formatTable(self, worksheet):
                # define range
                full_range = CellRange(min_col=worksheet.min_column,
                                       min_row=worksheet.min_row,
                                       max_col=worksheet.max_column,
                                       max_row=worksheet.max_row).coord

                # set table format
                mediumStyle =TableStyleInfo(name='TableStyleMedium1',
                                            showRowStripes=True)
                # create a table
                table = Table(ref=full_range,
                              displayName=(worksheet.title).replace(" ", "_"),
                              tableStyleInfo=mediumStyle)

                # add the table to the worksheet
                worksheet.add_table(table)

            def paste2Excel(self,worksheet,title):
                # title the worksheet
                worksheet.title = title

                # paste data frame into Excel worksheet
                for rows in dataframe_to_rows(self.data, index=False, header=True):
                    worksheet.append(rows)

                # format table from data
                self.formatTable(worksheet)

        # add virtual environment to path
        sys.path.insert(0, "src/.venv")
        
        # initialize log
        log = Log()

        # parse data from log
        log.parseData()

        # initialize "Dimensions" category
        res = Category(1)
        res(log)

        # initialize "Dimensions" category
        dim = Category(2)
        dim(log)

        # initialize "Contour Verify" category
        c_v = Category(3)
        c_v(log)

        # initialize "Corner" category
        cor = Category(4)
        cor(log)

        # open a workbook
        wb = Workbook()

        # create "Dimensions" sheet & paste data
        ws1 = wb.active
        res.paste2Excel(ws1,"Results")

        # create "Dimensions" sheet & paste data
        ws2 = wb.create_sheet()
        dim.paste2Excel(ws2,"Dimensions")


        # create "Contour Verify" sheet & paste data
        ws3 = wb.create_sheet()
        c_v.paste2Excel(ws3,"Contour Verify")

        # create "Corner" sheet & paste data
        ws4 = wb.create_sheet()
        cor.paste2Excel(ws4,"Corner")

        # save workbook
        wb.save(excel_file)

        # check if "dims.xlsx" was created from log
        if exists(excel_file):
            print("Log file successfully parsed to Excel!")
        else:
            print("ERROR: Log file unsuccessfully parsed to Excel...")

    def okay():
        len = len_entry.get()
        wid = wid_entry.get()
        hei = hei_entry.get()

        if not (type(len) == int or type(wid) == int or type(hei) == int):
            messagebox.showerror("Error", "Please input numeric tolerances")
            return

        filepath = import_entry.get()
        
        if filepath == '':
            messagebox.showerror("Error", "You must input a log file!")
            return
        elif not os.path.exists(filepath):
            messagebox.showerror("Error","Filepath does not exist...")
            return
        
        # check if excel file can be closed
        if not close_excel(excel_file):
            messagebox.showerror("Error","\"dims.xlsx\" is still open. You need to close it!")
            return

        script(filepath)
        window.destroy()
        exit(0)

    def skip():
        filepath = import_entry.get()

        if filepath == '':
            messagebox.showerror("Error", "You must input a log file!")
            return
        elif not os.path.exists(filepath):
            messagebox.showerror("Error","Filepath does not exist...")
            return
        
        # check if excel file can be closed
        if not close_excel(excel_file):
            messagebox.showerror("Error","\"dims.xlsx\" is still open. You need to close it!")
            return
        
        script(filepath)
        window.destroy()
        exit(0)

    def import_file():
        filepath = filedialog.askopenfilename(initialdir='/',title="Select a File",filetypes=[("Log files", ".txt .log")])
        
        import_entry.delete(0,"end")
        import_entry.insert(0,filepath)
        return
        
    def close_excel(excel_file):

        # remove if "dims.xlsx" exists
        if exists(excel_file):
            try:
                os.remove(excel_file)
                return True
            except OSError:  
                return False

    # specify excel file name
    excel_file = 'dims.xlsx'

    window = tk.Tk()
    window.title("Input Log Data")
    window.geometry("500x210")

    frame = tk.Frame(window)
    frame.pack()

    # Input Tolerance Frame
    input_tol_frame = tk.LabelFrame(frame,text="Input Tolerance")
    input_tol_frame.grid(row=0, column=0, sticky="ns", pady=10)

    len_label = tk.Label(input_tol_frame,text="Length")
    len_label.grid(row=0,column=0)
    len_entry = tk.Entry(input_tol_frame)
    len_entry.grid(row=1, column=0)

    wid_label = tk.Label(input_tol_frame,text="Width")
    wid_label.grid(row=0, column=1)
    wid_entry = tk.Entry(input_tol_frame)
    wid_entry.grid(row=1, column=1)

    hei_label = tk.Label(input_tol_frame,text="Height")
    hei_label.grid(row=0,column=2)
    hei_entry = tk.Entry(input_tol_frame)
    hei_entry.grid(row=1,column=2)

    # Export To Frame
    
    import_frame = tk.LabelFrame(frame, borderwidth=0)
    import_frame.grid(row=1, column=0)

    import_entry = tk.Entry(import_frame, width=50)
    import_entry.grid(row=0, column=1)
    import_button = tk.Button(import_frame, text="Import File", command=lambda:import_file())
    import_button.grid(row=0, column=0)

    # Confirm/Cancel Button
    validate_frame = tk.LabelFrame(frame, text="Run with or without tolerances?")
    validate_frame.grid(row=3, column=0)

    ok_button = tk.Button(validate_frame, text="With", width=20, command=okay)
    ok_button.grid(row=0, column=0)
    skip_button = tk.Button(validate_frame, text="Without", width=20, command=skip)
    skip_button.grid(row=0, column=1)

    # Adjust Padding
    for widget in input_tol_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    for widget in import_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    for widget in validate_frame.winfo_children():
        widget.grid_configure(padx=30, pady=10)

    window.mainloop()

if __name__ == "__main__":
    main()