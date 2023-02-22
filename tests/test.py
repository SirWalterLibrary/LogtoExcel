  
"""
with open("headers.txt") as file:
    print(len(file.readlines()))

"""

    

import openpyxl
from openpyxl.worksheet.cell_range import CellRange

# create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# populate some sample data    
worksheet["A1"] = "Fruit"
worksheet["B1"] = "Color"
worksheet["A2"] = "Apple"
worksheet["B2"] = "Red"
worksheet["A3"] = "Banana"
worksheet["B3"] = "Yellow"
worksheet["A4"] = "Coconut"
worksheet["B4"] = "Brown"

# define range
full_range = CellRange(min_col=worksheet.min_column,min_row=worksheet.min_row,max_col=worksheet.max_column,max_row=worksheet.max_row).coord

# define a table style
mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium1',
                                                      showRowStripes=True)
# create a table
table = openpyxl.worksheet.table.Table(ref=full_range,
                                       displayName='FruitColors',
                                       tableStyleInfo=mediumStyle)
# add the table to the worksheet
worksheet.add_table(table)

# save the workbook file
workbook.save('fruit.xlsx')